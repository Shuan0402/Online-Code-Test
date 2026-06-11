import csv
import io
from dataclasses import dataclass
from typing import List, Optional

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.security import SecurityManager
from app.models.enums import UserRole
from app.models.user import User
from app.services.candidate_password import generate_candidate_password
from app.services.tags import normalize_tags, replace_candidate_tags

REQUIRED_COLUMNS = ("真實姓名", "帳號")
SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}


@dataclass
class ParsedCandidateRow:
    row_number: int
    full_name: Optional[str]
    username: str


@dataclass
class BatchImportRowOutcome:
    row: int
    username: str
    full_name: Optional[str]
    status: str
    message: Optional[str] = None
    generated_password: Optional[str] = None


@dataclass
class BatchImportSummary:
    total: int
    created: int
    failed: int
    results: List[BatchImportRowOutcome]


class BatchImportFileError(Exception):
    pass


def _normalize_header(name: str) -> str:
    return (name or "").strip()


def _parse_csv(content: bytes) -> List[ParsedCandidateRow]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise BatchImportFileError("檔案缺少標題列")

    headers = {_normalize_header(h) for h in reader.fieldnames}
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise BatchImportFileError(f"缺少必要欄位：{', '.join(missing)}")

    rows: List[ParsedCandidateRow] = []
    for idx, record in enumerate(reader, start=2):
        full_name = (record.get("真實姓名") or "").strip() or None
        username = (record.get("帳號") or "").strip()
        if not username and not full_name:
            continue
        rows.append(ParsedCandidateRow(row_number=idx, full_name=full_name, username=username))
    return rows


def _parse_xlsx(content: bytes) -> List[ParsedCandidateRow]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    row_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(row_iter)
    except StopIteration:
        raise BatchImportFileError("檔案為空")

    headers = [_normalize_header(str(cell) if cell is not None else "") for cell in header_row]
    header_index = {name: i for i, name in enumerate(headers) if name}
    missing = [col for col in REQUIRED_COLUMNS if col not in header_index]
    if missing:
        raise BatchImportFileError(f"缺少必要欄位：{', '.join(missing)}")

    rows: List[ParsedCandidateRow] = []
    for row_number, cells in enumerate(row_iter, start=2):
        full_name_raw = cells[header_index["真實姓名"]] if header_index["真實姓名"] < len(cells) else None
        username_raw = cells[header_index["帳號"]] if header_index["帳號"] < len(cells) else None
        full_name = (str(full_name_raw).strip() if full_name_raw is not None else "") or None
        username = (str(username_raw).strip() if username_raw is not None else "")
        if not username and not full_name:
            continue
        rows.append(ParsedCandidateRow(row_number=row_number, full_name=full_name, username=username))
    return rows


def parse_candidate_upload(content: bytes, filename: str) -> List[ParsedCandidateRow]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        return _parse_csv(content)
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xls"):
        return _parse_xlsx(content)
    raise BatchImportFileError("僅支援 .csv、.xlsx、.xls 檔案格式")


def _validate_username(username: str) -> Optional[str]:
    if not username:
        return "帳號不可為空"
    if len(username) < 3:
        return "帳號至少需要 3 個字元"
    if len(username) > 50:
        return "帳號不可超過 50 個字元"
    return None


def batch_import_candidates(
    db: Session,
    rows: List[ParsedCandidateRow],
    tags: List[str],
) -> BatchImportSummary:
    normalized_tags = normalize_tags(tags)
    seen_usernames: set[str] = set()
    results: List[BatchImportRowOutcome] = []
    created = 0
    failed = 0

    for row in rows:
        username = row.username.strip()
        validation_error = _validate_username(username)
        if validation_error:
            failed += 1
            results.append(
                BatchImportRowOutcome(
                    row=row.row_number,
                    username=username,
                    full_name=row.full_name,
                    status="failed",
                    message=validation_error,
                )
            )
            continue

        if username in seen_usernames:
            failed += 1
            results.append(
                BatchImportRowOutcome(
                    row=row.row_number,
                    username=username,
                    full_name=row.full_name,
                    status="failed",
                    message="檔案內帳號重複",
                )
            )
            continue
        seen_usernames.add(username)

        existing = db.query(User).filter(User.username == username).first()
        if existing:
            failed += 1
            results.append(
                BatchImportRowOutcome(
                    row=row.row_number,
                    username=username,
                    full_name=row.full_name,
                    status="failed",
                    message="帳號已存在",
                )
            )
            continue

        plain_password = generate_candidate_password(username)
        new_user = User(
            username=username,
            full_name=row.full_name,
            password_hash=SecurityManager.hash_password(plain_password),
            role=UserRole.Candidate,
        )
        db.add(new_user)
        db.flush()

        if normalized_tags:
            replace_candidate_tags(db, new_user.id, normalized_tags)

        created += 1
        results.append(
            BatchImportRowOutcome(
                row=row.row_number,
                username=username,
                full_name=row.full_name,
                status="created",
                generated_password=plain_password,
            )
        )

    if created > 0:
        db.commit()
    else:
        db.rollback()

    return BatchImportSummary(
        total=len(rows),
        created=created,
        failed=failed,
        results=results,
    )
